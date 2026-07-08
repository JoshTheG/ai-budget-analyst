"""Park Maintenance Investment Priority Analysis - real PRNS data.

    python scripts/fetch_sanjose.py          # refresh the data (weekly)
    python scripts/park_conditions.py        # run the analysis

The real-world problem: San Jose PRNS maintains 200+ parks with limited
maintenance and capital dollars, and its capital funding (Construction &
Conveyance tax funds) is organized BY COUNCIL DISTRICT. Deciding where
the next dollar goes requires turning field-staff condition assessments
into a defensible, equity-aware priority list. That is what this script
produces, from the City's own weekly-published Park Condition Assessment
data (Survey123 scores by park and asset category, 2021-present).

Method - three transparent flags per park, no black box:
  condition flag   latest overall score below 70%
  declining flag   overall score down 5+ points since the park's first
                   assessment year
  equity flag      Healthy Places Index at or below the 25th percentile
                   (ActivateSJ guiding principle: equity & access)
Priority tier = number of flags (3 = act first). Ties sort by score.

Outputs (output/san_jose_parks/):
  park_priorities.xlsx   priority list, district rollup, asset systemic
                         view, citywide trend - styled workbook
  park_conditions_memo.md  written findings with every number computed
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook  # noqa: E402

from budget_analyst.report import HEADER_FILL, HEADER_FONT, _write_sheet  # noqa: E402

DATA = ROOT / "data" / "real" / "park_condition_assessment.csv"
OUT = ROOT / "output" / "san_jose_parks"

CONDITION_FLOOR = 70.0   # latest score below this -> condition flag
DECLINE_POINTS = 5.0     # drop of 5+ points since first year -> declining
EQUITY_PCTILE = 25.0     # HPI at/below this percentile -> equity flag

ASSET_COLS = {
    "PLAYGROUNDSCORETOTAL": "Playgrounds", "RESTROOMSCORETOTAL": "Restrooms",
    "COURTSCORETOTAL": "Courts", "FIELDSCORETOTAL": "Sports Fields",
    "TURFSCORETOTAL": "Turf", "PLANTSCORETOTAL": "Planting",
    "HARDSCAPESCORETOTAL": "Hardscape", "PKLOTSCORETOTAL": "Parking Lots",
    "PICNICSCORETOTAL": "Picnic Areas", "BENCHSCORETOTAL": "Benches",
    "DRINKFTNSCORETOTAL": "Drinking Fountains", "TREESSCORETOTAL": "Trees",
    "WASTERECSCORETOTAL": "Waste Receptacles",
    "WATERMGMTSCORETOTAL": "Water Management",
    "DOGPARKSCORETOTAL": "Dog Parks", "SKATEPARKSCORETOTAL": "Skate Parks",
    "EXERSTASCORETOTAL": "Exercise Stations",
}


def load_assessments() -> pd.DataFrame:
    """One assessment per park per year: prefer the annual Summer survey,
    then the most recent report, so Special follow-ups never double-count."""
    df = pd.read_csv(DATA)
    df["REPORTDATE"] = pd.to_datetime(df["REPORTDATE"], errors="coerce",
                                      format="mixed", utc=True)
    df["_summer"] = (df["ASSESSMENTTYPE"] == "Summer").astype(int)
    df = (df.sort_values(["_summer", "REPORTDATE"], ascending=False)
            .drop_duplicates(["PARKNAME", "ASSESSMENTYEAR"]))
    df["score"] = df["OVERALLSCORE"] * 100
    return df


def build_tables(df: pd.DataFrame) -> tuple[dict[str, pd.DataFrame], dict]:
    latest_year = int(df["ASSESSMENTYEAR"].max())
    latest = df[df["ASSESSMENTYEAR"] == latest_year].copy()

    # per-park first-vs-latest change
    firsts = (df.sort_values("ASSESSMENTYEAR")
                .drop_duplicates("PARKNAME")[["PARKNAME", "ASSESSMENTYEAR", "score"]]
                .rename(columns={"ASSESSMENTYEAR": "first_year",
                                 "score": "first_score"}))
    parks = latest.merge(firsts, on="PARKNAME", how="left")
    parks["change_pts"] = parks["score"] - parks["first_score"]

    parks["condition_flag"] = parks["score"] < CONDITION_FLOOR
    parks["declining_flag"] = parks["change_pts"] <= -DECLINE_POINTS
    parks["equity_flag"] = parks["HPIPCTILE"] <= EQUITY_PCTILE
    parks["priority_tier"] = (parks[["condition_flag", "declining_flag",
                                     "equity_flag"]].sum(axis=1))

    priorities = (parks.rename(columns={
        "PARKNAME": "park", "COUNCILDISTRICT": "council_district",
        "PARKTYPE": "park_type", "HPIPCTILE": "hpi_pctile",
        "score": "condition_score"})
        [["park", "council_district", "park_type", "condition_score",
          "first_year", "change_pts", "hpi_pctile", "condition_flag",
          "declining_flag", "equity_flag", "priority_tier"]]
        .sort_values(["priority_tier", "condition_score"],
                     ascending=[False, True])
        .round(2))

    # district rollup: where should district C&C dollars target?
    districts = (parks.groupby("COUNCILDISTRICT")
                 .agg(parks=("PARKNAME", "count"),
                      avg_score=("score", "mean"),
                      avg_change_pts=("change_pts", "mean"),
                      flagged_parks=("priority_tier", lambda s: int((s > 0).sum())),
                      tier3_parks=("priority_tier", lambda s: int((s == 3).sum())))
                 .reset_index()
                 .rename(columns={"COUNCILDISTRICT": "council_district"})
                 .sort_values("avg_score")
                 .round(2))

    # systemic asset view: which asset types are failing citywide?
    rows = []
    prior = df[df["ASSESSMENTYEAR"] == latest_year - 1]
    for col, label in ASSET_COLS.items():
        cur = latest[col].dropna() * 100
        if len(cur) < 10:
            continue
        prev = prior[col].dropna() * 100
        rows.append({"asset_category": label,
                     "parks_assessed": int(len(cur)),
                     "avg_score": round(float(cur.mean()), 2),
                     "pct_below_70": round(float((cur < 70).mean() * 100), 2),
                     "yoy_change_pts": round(float(cur.mean() - prev.mean()), 2)
                     if len(prev) >= 10 else None})
    assets = (pd.DataFrame(rows).sort_values("avg_score"))

    trend = (df.groupby("ASSESSMENTYEAR")
             .agg(parks_assessed=("PARKNAME", "count"),
                  avg_score=("score", "mean"))
             .reset_index()
             .rename(columns={"ASSESSMENTYEAR": "year"})
             .round(2))

    tier3 = priorities[priorities["priority_tier"] == 3]
    worst_asset = assets.iloc[0]
    worst_district = districts.iloc[0]
    facts = {
        "as_of": date.today().isoformat(),
        "latest_year": latest_year,
        "parks_assessed_latest": int(len(parks)),
        "years_covered": f"{int(df['ASSESSMENTYEAR'].min())}-{latest_year}",
        "citywide_avg_score": round(float(parks["score"].mean()), 1),
        "parks_below_floor": int(parks["condition_flag"].sum()),
        "parks_declining": int(parks["declining_flag"].sum()),
        "parks_equity_flag": int(parks["equity_flag"].sum()),
        "tier3_count": int(len(tier3)),
        "tier3_parks": ", ".join(tier3["park"].head(12)),
        "worst_asset": str(worst_asset["asset_category"]),
        "worst_asset_score": float(worst_asset["avg_score"]),
        "worst_asset_below70_pct": float(worst_asset["pct_below_70"]),
        "worst_district": int(worst_district["council_district"]),
        "worst_district_score": float(worst_district["avg_score"]),
        "best_district": int(districts.iloc[-1]["council_district"]),
        "best_district_score": float(districts.iloc[-1]["avg_score"]),
    }
    tables = {"park_priorities": priorities, "district_rollup": districts,
              "asset_categories": assets, "citywide_trend": trend}
    return tables, facts


def write_memo(facts: dict, out_dir: Path) -> Path:
    f = facts
    memo = f"""# Park Maintenance Investment Priority Analysis

**Source:** City of San Jose Park Condition Assessment (open data,
published weekly) · {f['years_covered']} · analyzed {f['as_of']}

## Summary

In the {f['latest_year']} assessment cycle, {f['parks_assessed_latest']}
parks were scored; the citywide average condition is
{f['citywide_avg_score']}/100. {f['parks_below_floor']} parks score below
the 70-point condition floor, {f['parks_declining']} have declined 5+
points since their first assessment, and {f['parks_equity_flag']} sit in
the lowest quartile of the Healthy Places Index (the ActivateSJ equity
lens).

## Priority findings

- **{f['tier3_count']} parks carry all three flags** (poor + declining +
  high-equity-need) and are the strongest candidates for the next
  maintenance/C&C dollar: {f['tier3_parks']}.
- **Systemic asset gap:** {f['worst_asset']} is the weakest category
  citywide (avg {f['worst_asset_score']}/100;
  {f['worst_asset_below70_pct']}% of parks below 70) - a program-level
  funding case, not a park-by-park one.
- **District picture:** Council District {f['worst_district']} averages
  {f['worst_district_score']}/100 vs. District {f['best_district']} at
  {f['best_district_score']}/100 - relevant to targeting the
  district-based Construction & Conveyance tax funds.

## Method (fully transparent)

One assessment per park-year (annual Summer survey preferred). Three
boolean flags: condition (<70/100), declining (-5 pts or more since first
year), equity (HPI ≤ 25th percentile). Priority tier = flag count. Every
figure is computed in `scripts/park_conditions.py`; no modeling, no
weights to argue about.

## Recommended next steps

1. Validate tier-3 parks against current CIP project list and C&C fund
   balances by district (this toolkit's `fund_summary` view).
2. Cost the {f['worst_asset'].lower()} gap as a single program line for
   the next budget development cycle.
3. Re-run weekly as new assessments post; scores move when crews do.
"""
    path = out_dir / "park_conditions_memo.md"
    path.write_text(memo, encoding="utf-8")
    return path


def main() -> int:
    if not DATA.exists():
        print("data/real/park_condition_assessment.csv missing - run "
              "scripts/fetch_sanjose.py first")
        return 1
    OUT.mkdir(parents=True, exist_ok=True)
    df = load_assessments()
    tables, facts = build_tables(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Park Maintenance Investment Priorities"
    from openpyxl.styles import Font
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = (f"San Jose Park Condition Assessment {facts['years_covered']}"
                f" · generated {facts['as_of']}")
    ws["A2"].font = Font(italic=True, size=9)
    ws.append([])
    ws.append(["Metric", "Value"])
    for cell in ws[4]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for k, v in facts.items():
        ws.append([k, v])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 60
    for name, t in tables.items():
        _write_sheet(wb, name, t)
    xlsx = OUT / "park_priorities.xlsx"
    wb.save(xlsx)
    memo = write_memo(facts, OUT)

    print(f"analyzed {facts['parks_assessed_latest']} parks "
          f"({facts['years_covered']})")
    print(f"  citywide avg: {facts['citywide_avg_score']}/100 | "
          f"below-70: {facts['parks_below_floor']} | "
          f"declining: {facts['parks_declining']} | "
          f"tier-3 priorities: {facts['tier3_count']}")
    print(f"wrote {xlsx}")
    print(f"wrote {memo}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
