"""Excel workbook + markdown memo output.

The workbook is meant to look like something a budget office would
circulate: styled headers, currency formats, red/green variance
highlighting, an embedded budget-vs-actual chart, and a Summary sheet
with the headline figures and a generation timestamp.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
OVER_FILL = PatternFill("solid", fgColor="F8CBAD")    # overruns / negatives
UNDER_FILL = PatternFill("solid", fgColor="C6EFCE")   # healthy balances
MONEY_COLS = {"budget", "actual", "variance", "total", "change", "encumbrance",
              "available", "appropriation", "expended", "encumbered",
              "revenue_target", "revenue_actual", "revenue_variance",
              "revenue_estimate", "net_activity"}
PCT_COLS = {"variance_pct", "pct_spent", "pct_committed", "attainment_pct",
            "change_pct"}
# columns where red should flag values above/below a threshold
_RED_IF_NEGATIVE = {"available", "variance", "net_activity"}
_RED_IF_OVER_100 = {"pct_spent", "pct_committed"}


def _write_sheet(wb, name: str, df) -> None:
    ws = wb.create_sheet(name[:31])
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in df.itertuples(index=False):
        ws.append(list(row))
    n_rows = len(df) + 1
    for i, col in enumerate(df.columns, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = max(len(str(col)) + 2, 14)
        col_l = str(col).lower()
        if col_l in MONEY_COLS:
            for cell in ws[letter][1:]:
                cell.number_format = '#,##0.00'
        elif col_l in PCT_COLS:
            for cell in ws[letter][1:]:
                cell.number_format = '0.00"%"'
        rng = f"{letter}2:{letter}{n_rows}"
        if col_l in _RED_IF_NEGATIVE:
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="lessThan", formula=["0"], fill=OVER_FILL))
        elif col_l in _RED_IF_OVER_100:
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="greaterThan", formula=["100"], fill=OVER_FILL))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="lessThanOrEqual", formula=["100"], fill=UNDER_FILL))
    ws.freeze_panes = "A2"


def _add_variance_chart(wb, df, entity_col: str) -> None:
    """Embed a clustered budget-vs-actual bar chart on the variance sheet."""
    if "variance_by_entity" not in wb.sheetnames:
        return
    ws = wb["variance_by_entity"]
    cols = list(df.columns)
    try:
        b_idx = cols.index("budget") + 1
        a_idx = cols.index("actual") + 1
    except ValueError:
        return
    chart = BarChart()
    chart.type, chart.style = "col", 10
    chart.title = f"Budget vs. Actual by {entity_col} (latest period)"
    chart.y_axis.title = "Dollars"
    n = len(df) + 1
    for idx, label in ((b_idx, "Budget"), (a_idx, "Actual")):
        ref = Reference(ws, min_col=idx, min_row=1, max_row=n)
        chart.add_data(ref, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n))
    chart.width, chart.height = 24, 10
    ws.add_chart(chart, f"A{n + 3}")


_SUMMARY_LABELS = {
    "total_budget": "Total budget (latest period)",
    "total_actual": "Total actual expenditures",
    "total_variance": "Net variance (budget - actual)",
    "overall_pct_spent": "% of budget spent",
    "total_encumbrance": "Total encumbered",
    "total_available": "Total available balance",
    "overall_pct_committed": "% committed (spent + encumbered)",
    "total_revenue_target": "Revenue target",
    "total_revenue_actual": "Revenue collected",
    "revenue_attainment_pct": "Revenue attainment %",
    "forecast_next_period": "Next-period projection",
    "anomaly_count": "Variance outliers flagged",
}


def write_outputs(result: dict, memo: str, mapping: dict, out_dir: str) -> dict:
    """Write workbook, memo, and an audit trail of the schema mapping."""
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Budget Analysis Summary"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = f"Generated {datetime.now():%Y-%m-%d %H:%M} by AI Budget Analyst"
    ws["A2"].font = Font(italic=True, size=9)
    ws.append([])
    ws.append(["Metric", "Value"])
    for cell in ws[4]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    facts = result["facts"]
    ordered = [(k, facts[k]) for k in _SUMMARY_LABELS if k in facts]
    ordered += [(k, v) for k, v in facts.items() if k not in _SUMMARY_LABELS]
    for k, v in ordered:
        ws.append([_SUMMARY_LABELS.get(k, k), v])
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 40
    for name, df in result["tables"].items():
        _write_sheet(wb, name, df)
    if "variance_by_entity" in result["tables"]:
        _add_variance_chart(wb, result["tables"]["variance_by_entity"],
                            facts.get("entity_column", "entity"))

    xlsx = out / "budget_analysis.xlsx"
    wb.save(xlsx)

    memo_path = out / "budget_memo.md"
    memo_path.write_text(memo, encoding="utf-8")

    audit = out / "schema_mapping.json"
    audit.write_text(json.dumps(mapping, indent=2), encoding="utf-8")

    return {"workbook": str(xlsx), "memo": str(memo_path), "mapping": str(audit)}
